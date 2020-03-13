"""
Converts the graph theory data in source space
__author__: Amna Hyder
__date__: August 2018
"""

import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
# Global Variables
#change conditions based on folders
path = os.getcwd()
conditions = ['Group1', 'Group2']

global_headings = ['Subject','Density','Global Efficiency','Modularity','Clustering Coeff', 'Condition']
electrode_headings = ['Subject','Degree','Clustering Coefficient','Local Reachability','Betweenness','Hub Value','Authority', 'Condition', 'Electrode' ]
electrodes = ['F9','A1','P9','Fp1','F7','T7','P7','O1','F3','C3','P3','Fpz','Fz','Cz','Pz','Oz','F4','C4','P4','Fp2','F8','T8','P8','O2','F10','A2','P10']

def xls_to_csv(sh, csv_name):
    with open( csv_name +'.csv', 'wb') as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])

def isnotDS(f):
    return not f.startswith('.')

def header_sheet(ws, condition):
    subjects = os.listdir(os.getcwd())
    #NotADirectoryError: [Errno 20] Not a directory: '.DS_Store/.DS_Store_g_quantitative.csv'
    #subjects = list(filter(lambda f: f is not '.DS_Store', subjects))
    subjects = list(filter(isnotDS, subjects))
    for i in range(len(subjects)):
        curr_sub = subjects[i]
        wb_name = curr_sub + '/'  + curr_sub +  '_g_quantitative.csv'
        # Get data from the csv named above
        with open(wb_name) as fd:
            reader=csv.reader(fd)
            for i, row in enumerate(reader):
                #  The global data is in the third row (second index)
                if i == 2:
                    # Add this row to the workbook
                    ws.append([curr_sub] + row[:-1]+ [condition])
    #Leave an empty row between sets
    #ws.append([])

def electrode_sheet(wb, c, ws_e):
    # Set electrode data in its own sheet
    subjects = os.listdir(os.getcwd())

    for e in range(len(electrodes)):
        # Create a worksheet for this electrode
        curr_electrode = electrodes[e]
        # Comment out line below if you don't want condition repeated above
        # this set of data
        # Add a row for the heading names
        for i in range(len(subjects)):
            curr_sub = subjects[i]
            wb_name = curr_sub + '/' + curr_sub + '_g_quantitative.csv'
            # Open the workbook named above
            with open(wb_name) as fd:
                reader=csv.reader(fd)
                for i, row in enumerate(reader):
                    # Starting from 7 rows down (where the data begins)
                    # Add the data in this (i-th) row to the current worksheet
                    if i == e + 6:
                        num_in_row = [float(num) for num in row[:-1]]
                        ws_e.append([curr_sub] + row[:-1] + [c, curr_electrode])
                        #print([curr_sub] + row + [condition, curr_electrode])
                        #print([curr_sub])
                        #print(row[:-1])
                        #print([condition, curr_electrode])



# Create a separate workbook for every condition
gwb = Workbook()
gws = gwb.active

# Insert all the header data FOR GLOBAL MEASURES
gws.append(global_headings)
for condition in conditions:
    os.chdir(condition)
    header_sheet(gws, condition)
    os.chdir('..')

# Insert all electrode data
ws_e =  gwb.create_sheet('all_data')
#ws_e = gwb['all_data']
ws_e.append(electrode_headings )
for condition in conditions:
    os.chdir(condition)
    electrode_sheet(gwb, condition, ws_e)
    os.chdir('..')
gwb.save('ALL_Data' + '.xlsx')


xls_to_csv(gws, 'global')
xls_to_csv(ws_e, 'electrode')




