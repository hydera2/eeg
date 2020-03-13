import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
# Global Variables
#change conditions based on folders
path = os.getcwd()
conditions = ['A', 'B']

global_headings = ['Subject','Density','Global Efficiency','Modularity','Clustering Coeff', 'Condition']
electrode_headings = ['Subject','Degree','Clustering Coefficient','Local Reachability','Betweenness','Hub Value','Authority', 'Condition', 'Electrode' ]
electrodes = ['TAL','TPL','FL','CL','PL','FpM','FM','CM','PM','OpM','FR','CR','PR','TAR','TPR']
def isnotDS(f):
    return f is not '.DS_Store'

def header_sheet(ws, condition):
    subjects = os.listdir(os.getcwd())
    #NotADirectoryError: [Errno 20] Not a directory: '.DS_Store/.DS_Store_g_quantitative.csv'
    #subjects = list(filter(lambda f: f is not '.DS_Store', subjects))
    subjects = list(filter(isnotDS, subjects))
    for i in range(len(subjects)):
        curr_sub = subjects[i]
        print(curr_sub)
        wb_name = curr_sub 
        # Get data from the csv named above
        with open(wb_name) as fd:
            reader=csv.reader(fd)
            for i, row in enumerate(reader):
                #  The global data is in the third row (second index)
                if i == 2:
                    # Add this row to the workbook
                    ws.append([curr_sub] + row[:-1]+ [condition])
    #Leave an empty row between sets
    #ws.append([])

def electrode_sheet(wb, c, ws_e):
    # Set electrode data in its own sheet
    subjects = os.listdir(os.getcwd())

    for e in range(len(electrodes)):
        # Create a worksheet for this electrode
        curr_electrode = electrodes[e]
        # Comment out line below if you don't want condition repeated above
        # this set of data
        # Add a row for the heading names
        for i in range(len(subjects)):
            curr_sub = subjects[i]
            wb_name = curr_sub
            # Open the workbook named above
            with open(wb_name) as fd:
                reader=csv.reader(fd)
                for i, row in enumerate(reader):
                    # Starting from 7 rows down (where the data begins)
                    # Add the data in this (i-th) row to the current worksheet
                    if i == e + 6:
                        num_in_row = [float(num) for num in row[:-1]]
                        ws_e.append([curr_sub] + row[:-1] + [c, curr_electrode])
                        #print([curr_sub] + row + [condition, curr_electrode])
                        #print([curr_sub])
                        #print(row[:-1])
                        #print([condition, curr_electrode])



# Create a separate workbook for every condition
gwb = Workbook()
gws = gwb.active

# Insert all the header data FOR GLOBAL MEASURES
gws.append(global_headings)
for condition in conditions:
    os.chdir(condition)
    header_sheet(gws, condition)
    os.chdir('..')

# Insert all electrode data
ws_e =  gwb.create_sheet('all_data')
#ws_e = gwb['all_data']
ws_e.append(electrode_headings )
for condition in conditions:
    os.chdir(condition)
    electrode_sheet(gwb, condition, ws_e)
    os.chdir('..')
gwb.save('ALL_Data' + '.xlsx')
